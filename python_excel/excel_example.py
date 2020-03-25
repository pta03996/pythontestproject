import openpyxl as xl
from openpyxl.chart import BarChart, Reference

workbook = xl.load_workbook('transactions.xlsx')
workbook_sheet = workbook['Sheet1']
### diff code but same result
# workbook_cell = workbook_sheet['a1']
# workbook_cell = workbook_sheet.cell(1,1)
# get how many row in a sheet
print(workbook_sheet.max_row)

for row in range(2, workbook_sheet.max_row + 1):
    cell = workbook_sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = workbook_sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

value = Reference(workbook_sheet, min_col=4, min_row=2, max_col=4, max_row=workbook_sheet.max_row)
chart = BarChart()
chart.add_data(value)
workbook_sheet.add_chart(chart, 'E3')
workbook.save('transaction2.xlsx')
